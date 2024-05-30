import io
from typing import Dict, List, Union

import aiosqlite
import pandas as pd
import prettytable as pt
from aiosqlite import Connection
from openpyxl import load_workbook
from resources import config
import resources.config


class ExcelParser:
    @staticmethod
    def parse_excel(excel_bytes: bytes) -> Dict[str, pd.DataFrame]:
        """
        Парсинг данных из файла Excel и возврат словаря с DataFrame для каждого листа.

        Args:
            excel_bytes (bytes): Двоичные данные файла Excel.

        Returns:
            Dict[str, pd.DataFrame]: Словарь, где ключ - название листа, значение - DataFrame с данными.
        """
        data = {}
        wb = load_workbook(filename=io.BytesIO(excel_bytes))
        required_columns = [
            "day_of_week",
            "start_time",
            "end_time",
            "location",
            "subject_name",
            "email",
            "last_name",
            "first_name",
        ]

        for sheet in wb.sheetnames:
            df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=sheet)
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(
                    f"Отсутствует столбец  '{sheet}': {', '.join(missing_columns)}"
                )
            df["start_time"] = pd.to_datetime(
                df["start_time"], format="%H:%M:%S"
            ).dt.time
            df["end_time"] = pd.to_datetime(df["end_time"], format="%H:%M:%S").dt.time
            data[sheet] = df
        return data


class ScheduleDB:
    @staticmethod
    async def connect_to_db() -> Connection:
        """
        Установление соединения с базой данных.

        Returns:
            Connection: Объект соединения с базой данных.
        """
        return await aiosqlite.connect(config.get_name_db())

    async def create_tables(self) -> None:
        """
        Создание таблиц в базе данных.
        """
        conn = await self.connect_to_db()
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS Groups (
                group_id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_name VARCHAR(50)
            );
            """
        )
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS Schedule (
                schedule_id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INT,
                day_of_week VARCHAR(20),
                FOREIGN KEY (group_id) REFERENCES Groups(group_id)
            );
            """
        )
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS Subjects (
                subject_id INTEGER PRIMARY KEY AUTOINCREMENT,
                schedule_id INT,
                subject_name VARCHAR(100),
                start_time TIME,
                end_time TIME,
                location VARCHAR(100),
                FOREIGN KEY (schedule_id) REFERENCES Schedule(schedule_id)
            );
            """
        )
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS Emails (
                email_id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INT,
                email VARCHAR(100),
                last_name VARCHAR(100),
                first_name VARCHAR(100),
                FOREIGN KEY (group_id) REFERENCES Groups(group_id)
            );
            """
        )
        await conn.commit()
        await conn.close()
        print("Tables created successfully.")

    async def clear_database(self) -> None:
        """
        Очистка базы данных (удаление всех таблиц).
        """
        conn = await self.connect_to_db()
        await conn.execute("DROP TABLE IF EXISTS Subjects")
        await conn.execute("DROP TABLE IF EXISTS Schedule")
        await conn.execute("DROP TABLE IF EXISTS Groups")
        await conn.execute("DROP TABLE IF EXISTS Emails")
        await conn.commit()
        await conn.close()

    async def get_group(self, group_name: str) -> Union[int, None]:
        """
        Получение идентификатора группы по её названию.

        Args:
            group_name (str): Название группы.

        Returns:
            Union[int, None]: Идентификатор группы или None, если группа не найдена.
        """
        conn = await self.connect_to_db()
        cursor = await conn.execute(
            "SELECT group_id FROM Groups WHERE LOWER(group_name) = ?",
            (group_name.lower(),),
        )
        group = await cursor.fetchone()
        await conn.close()
        return group[0] if group else None

    async def create_group(self, group_name: str) -> int:
        """
        Создание новой группы в базе данных.

        Args:
            group_name (str): Название группы.

        Returns:
            int: Идентификатор новой группы.
        """
        conn = await self.connect_to_db()
        cursor = await conn.execute(
            "INSERT INTO Groups (group_name) VALUES (LOWER(?))", (group_name,)
        )
        await conn.commit()
        last_row_id = cursor.lastrowid
        await conn.close()
        return last_row_id

    async def get_all_groups(self) -> List[str]:
        """
        Получение всех групп из базы данных.
        """
        conn = await self.connect_to_db()
        cursor = await conn.execute("SELECT group_name FROM Groups")
        groups = await cursor.fetchall()
        await conn.close()
        return [group[0] for group in groups]

    async def get_schedule(
        self, group_id: int, day_of_week: str
    ) -> Union[List[int], None]:
        """
        Получение идентификаторов расписаний для определенной группы и дня недели.

        Args:
            group_id (int): Идентификатор группы.
            day_of_week (str): День недели.

        Returns:
            Union[List[int], None]: Список идентификаторов расписаний или None, если расписание не найдено.
        """
        conn = await self.connect_to_db()
        cursor = await conn.execute(
            "SELECT schedule_id FROM Schedule WHERE group_id = ? AND day_of_week = ?",
            (group_id, day_of_week),
        )
        schedule = await cursor.fetchall()
        await conn.close()
        return [row[0] for row in schedule] if schedule else None

    async def create_schedule(self, group_id: int, day_of_week: str) -> int:
        """
        Создание нового расписания для определенной группы и дня недели.

        Args:
            group_id (int): Идентификатор группы.
            day_of_week (str): День недели.

        Returns:
            int: Идентификатор нового расписания.
        """
        conn = await self.connect_to_db()
        cursor = await conn.execute(
            "INSERT INTO Schedule (group_id, day_of_week) VALUES (?, ?)",
            (group_id, day_of_week),
        )
        await conn.commit()
        last_row_id = cursor.lastrowid
        await conn.close()
        return last_row_id

    async def add_subject(
        self,
        schedule_id: int,
        subject_name: str,
        start_time: str,
        end_time: str,
        location: str,
    ) -> int:
        """
        Добавление нового предмета в базу данных.

        Args:
            schedule_id (int): Идентификатор расписания.
            subject_name (str): Название предмета.
            start_time (str): Время начала занятия.
            end_time (str): Время окончания занятия.
            location (str): Аудитория.

        Returns:
            int: Идентификатор нового предмета.
        """
        conn = await self.connect_to_db()
        cursor = await conn.execute(
            """
                INSERT INTO Subjects (schedule_id, subject_name, start_time, end_time, location) 
                VALUES (?, ?, ?, ?, ?)
            """,
            (schedule_id, subject_name, start_time, end_time, location),
        )
        await conn.commit()
        last_row_id = cursor.lastrowid
        await conn.close()
        return last_row_id

    async def add_email(
        self,
        group_id: int,
        email: str,
        last_name: str,
        first_name: str,
    ) -> int:
        """
        Добавление нового email в базу данных.

        Args:
            group_id (int): Идентификатор группы.
            email (str): Email адрес.
            last_name (str): Фамилия.
            first_name (str): Имя.

        Returns:
            int: Идентификатор нового email.
        """
        conn = await self.connect_to_db()
        cursor = await conn.execute(
            """
                INSERT INTO Emails (group_id, email, last_name, first_name) 
                VALUES (?, ?, ?, ?)
            """,
            (group_id, email, last_name, first_name),
        )
        await conn.commit()
        last_row_id = cursor.lastrowid
        await conn.close()
        return last_row_id

    async def get_emails_by_group(self, group_name: str) -> Union[List[str], None]:
        """
        Получение email адресов по названию группы.

        Args:
            group_name (str): Название группы.

        Returns:
            Union[List[str], None]: Список email адресов или None, если email не найдены.
        """
        group_id = await self.get_group(group_name)
        if group_id is None:
            return None

        conn = await self.connect_to_db()
        cursor = await conn.execute(
            "SELECT email FROM Emails WHERE group_id = ?", (group_id,)
        )
        emails = await cursor.fetchall()
        await conn.close()
        return [row[0] for row in emails] if emails else None

    async def get_weekly_schedule_by_group(self, group_name: str) -> str:
        """
        Получение недельного расписания для определенной группы.

        Args:
            group_name (str): Название группы.

        Returns:
            str: Текстовое представление недельного расписания в виде таблицы.
        """
        group_id = await self.get_group(group_name)
        if group_id is None:
            return "Такой группы не существует, введите другую"
        days_of_week = [
            "Понедельник",
            "Вторник",
            "Среда",
            "Четверг",
            "Пятница",
            "Суббота",
        ]
        weekly_schedule = {}
        conn = await self.connect_to_db()

        for day in days_of_week:
            schedule_ids = await self.get_schedule(group_id, day)
            if schedule_ids is not None:
                for schedule_id in schedule_ids:
                    cursor = await conn.execute(
                        """
                        SELECT subject_name, start_time, end_time, location
                        FROM Subjects
                        WHERE schedule_id = ?
                        """,
                        (schedule_id,),
                    )
                    schedule = await cursor.fetchall()
                    if day not in weekly_schedule:
                        weekly_schedule[day] = []
                    weekly_schedule[day].extend(schedule)
        await conn.close()

        sorted_schedule = {}
        for day in days_of_week:
            if day in weekly_schedule:
                sorted_schedule[day] = sorted(weekly_schedule[day], key=lambda x: x[1])

        table = pt.PrettyTable()
        table.field_names = ["День", "Предмет", "Начало", "Конец", "Аудитория"]
        for day, classes in sorted_schedule.items():
            for subject, start, end, room in classes:
                table.add_row([day, subject, start, end, room])

        return f"```Расписание\n{table}\n```"

    async def insert_data_from_excel(self, excel_bytes: bytes) -> None:
        """
        Вставка данных из файла Excel в базу данных.

        Args:
            excel_bytes (bytes): Двоичные данные файла Excel.
        """
        await self.clear_database()
        await self.create_tables()
        data = ExcelParser.parse_excel(excel_bytes)
        conn = await self.connect_to_db()

        for group_name, df in data.items():
            cursor = await conn.execute(
                "INSERT INTO Groups (group_name) VALUES (?)", (group_name,)
            )
            group_id = cursor.lastrowid

            # Вставка данных в таблицу Emails
            for _, row in df.iterrows():
                email = row["email"]
                last_name = row["last_name"]
                first_name = row["first_name"]
                await conn.execute(
                    """
                    INSERT INTO Emails (group_id, email, last_name, first_name) 
                    VALUES (?, ?, ?, ?)
                    """,
                    (group_id, email, last_name, first_name),
                )

            for index, row in df.iterrows():
                # Вставка данных в таблицу Schedule
                day_of_week = row["day_of_week"]
                cursor = await conn.execute(
                    "INSERT INTO Schedule (group_id, day_of_week) VALUES (?, ?)",
                    (group_id, day_of_week),
                )
                schedule_id = cursor.lastrowid

                # Вставка данных в таблицу Subjects
                start_time = row["start_time"].strftime("%H:%M:%S")
                end_time = row["end_time"].strftime("%H:%M:%S")
                location = row["location"]
                subject_name = row["subject_name"]
                await conn.execute(
                    """
                    INSERT INTO Subjects (schedule_id, subject_name, start_time, end_time, location) 
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (schedule_id, subject_name, start_time, end_time, location),
                )

        await conn.commit()
        await conn.close()


db = ScheduleDB()
